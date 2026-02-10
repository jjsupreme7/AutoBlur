"""Excel parser using openpyxl."""

import re
from typing import List

import openpyxl

from .base import BaseParser, TextRegion, RedactionTarget


class ExcelParser(BaseParser):
    def extract(self) -> List[TextRegion]:
        wb = openpyxl.load_workbook(self.file_path)
        regions = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows(min_row=1):
                for cell in row:
                    if cell.value is not None and str(cell.value).strip():
                        regions.append(TextRegion(
                            text=str(cell.value),
                            location={
                                'sheet': sheet_name,
                                'row': cell.row,
                                'col': cell.column,
                                'col_letter': cell.column_letter,
                            },
                            source_file=self.file_path,
                        ))
        wb.close()
        return regions

    def redact(self, targets: List[RedactionTarget], output_path: str) -> None:
        wb = openpyxl.load_workbook(self.file_path)
        for target in targets:
            loc = target.region.location
            ws = wb[loc['sheet']]
            cell = ws.cell(row=loc['row'], column=loc['col'])
            original = str(cell.value)
            # Replace the matched text within the cell
            if target.region.text == original:
                cell.value = target.replacement
            else:
                cell.value = original.replace(target.region.text, target.replacement)
        wb.save(output_path)
        wb.close()
